from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from io import BytesIO


def excel_to_pdf(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    output = BytesIO()
    c = canvas.Canvas(output, pagesize=landscape(letter))

    # Get a sample style and modify it
    style = getSampleStyleSheet()["Normal"]
    style.wordWrap = 'CJK'

    # Calculate column widths
    col_widths = []
    for col in ws.iter_cols():
        max_width = 0
        for cell in col:
            text = str(cell.value)
            wrapped_text = Paragraph(text, style)
            w, _ = wrapped_text.wrap(100, 20)
            max_width = max(max_width, w)
        col_widths.append(max_width + 10)  # Add some padding

    # Calculate row heights
    row_heights = []
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            text = str(cell.value)
            wrapped_text = Paragraph(text, style)
            _, h = wrapped_text.wrap(100, 20)
            max_height = max(max_height, h)
        row_heights.append(max_height + 5)  # Add some padding

    # Draw the content on the PDF canvas
    y = 750
    for row_index, row in enumerate(ws.iter_rows()):
        x = 50
        for col_index, cell in enumerate(row):
            text = str(cell.value)
            wrapped_text = Paragraph(text, style)
            wrapped_text.wrap(col_widths[col_index], row_heights[row_index])
            wrapped_text.drawOn(c, x, y)
            x += col_widths[col_index]
        y -= row_heights[row_index]

    c.showPage()
    c.save()
    output.seek(0)

    return output


# new version
# implementation with styles.
"""
from io import BytesIO
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


def excel_to_pdf(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))

    data = []

    styles = getSampleStyleSheet()
    style = styles["BodyText"]

    for row in ws.iter_rows():
        data_row = []
        for cell in row:
            text = str(cell.value)
            wrapped_text = Paragraph(text, style)
            data_row.append(wrapped_text)
        data.append(data_row)

    table = Table(data)
    table.setStyle(TableStyle([
        # ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    doc.build([table])
    output.seek(0)

    return output

"""
